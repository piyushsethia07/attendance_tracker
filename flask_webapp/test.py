from sqlalchemy import create_engine
from flask_app import Class, ClassSchedule, Teacher, Student, StudentClass, Module
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

selected_columns_student = ['Student ID',
                                      'AC/IT',
                                        'Att',
                                  'Last Name',
                                 'First Name',
                             'Preferred name']

def add_data_after_reading_csv_file(excel_file_path):
    wb = load_workbook(excel_file_path, data_only=True)  # Use data_only=True to read evaluated values

    # Create an empty list to store the data
    data = []
    
    # Iterate through all sheets in the workbook
    for sheet in wb.sheetnames:
        # Read each sheet as a DataFrame
        df = pd.read_excel(excel_file_path, sheet_name=sheet, header=None)
        df = df.dropna(how='all')
        data.append(df)

    for df in data:
        if 'date' in df.iloc[0][0].lower():
            metadata = {}
            for index in range(len(df.iloc[0])):
                if pd.notna(df.iloc[0][index]):
                    try:
                        key = df.iloc[0][index].split(':')[0].strip().lower()
                        if key.lower() == 'class total':
                            value = df.iloc[0][index+1]
                        else:
                            value = df.iloc[0][index].split(':')[1].strip()
                        metadata[key] = value
                    except:
                        pass
                        # print('Invalid data format')
            temp_df = df
            temp_df = temp_df.drop(0)
            
            # Set the second row as column names
            temp_df.columns = temp_df.iloc[0]
            
            # Reset the index
            temp_df = temp_df.reset_index(drop=True)
            
            temp_df = temp_df.drop(0)
            return(metadata, temp_df)

# Create a SQLAlchemy engine and session
engine = create_engine('sqlite:///myapp.db')  # Replace with your SQLite database URI
Session = sessionmaker(bind=engine)
session = Session()


def add_class(data):
    class_record = session.query(Class).filter_by(class_name=data['class']).first()
    if not class_record:
        new_class = Class(class_name=data['class'])
        session.add(new_class)
        session.commit()
        return f"{data['class']} added successfully"
    else:
        return f"Class with {data['class']}  already exists"

# Function to add data to the Teacher table

def add_teacher(data):
    new_teacher = session.query(Teacher).filter_by(teacher_name=data['trainer']).first()
    if not new_teacher:
        new_teacher = Teacher(teacher_name=data['trainer'])
        session.add(new_teacher)
        session.commit()
        return f"{data['trainer']} added successfully"
    else:
        return f"Teacher with {data['trainer']}  already exists"
    
def add_module(data):
    new_teacher = session.query(Module).filter_by(module_name=data['module']).first()
    if not new_teacher:
        new_teacher = Module(module_name=data['module'])
        session.add(new_teacher)
        session.commit()
        return f"{data['module']} added successfully"
    else:
        return f"Module with {data['module']}  already exists"

# Function to add data to the ClassSchedule table
def add_class_schedule(data):
    class_name = data['class']
    trainer_name = data['trainer']
    parsed_date = datetime.strptime(data['date'], '%m/%d/%Y').date()
    # Query the Class and Teacher tables to get their corresponding IDs
    class_record = session.query(Class).filter_by(class_name=class_name).first()
    teacher_record = session.query(Teacher).filter_by(teacher_name=trainer_name).first()

    if class_record and teacher_record:
        new_class_schedule = ClassSchedule(
            date=parsed_date,
            module=data['module'],
            class_name=class_record.class_name,
            trainer_name=teacher_record.teacher_name,
            class_total=data['class total']
        )
        session.add(new_class_schedule)
        session.commit()
        return "Success"
    else:
        add_class(data)
        add_teacher(data)
        add_module(data)
        class_record = session.query(Class).filter_by(class_name=class_name).first()
        teacher_record = session.query(Teacher).filter_by(teacher_name=trainer_name).first()
        new_class_schedule = ClassSchedule(
            date=parsed_date,
            module=data['module'],
            class_name=class_record.class_name,
            trainer_name=teacher_record.teacher_name,
            class_total=data['class total']
        )
        session.add(new_class_schedule)
        session.commit()
        return ("Class or Teacher not found in the database. Added to database and then class schedule updated")

def add_student(df):
    # Loop through the rows of your DataFrame and insert or update each row in the 'student' table
    for index, row in df.iterrows():
        existing_student = session.query(Student).filter_by(student_id=str(row['Student ID'])).first()
        
        if existing_student:
            # Student with the same student_id already exists, update the existing record
            existing_student.course_enrolled = row['AC/IT']
            existing_student.attendance = row['Att']
            existing_student.last_name = row['Last Name']
            existing_student.first_name = row['First Name']
            existing_student.preferred_name = row['Preferred name']
        else:
            # Student with this student_id doesn't exist, create a new record
            student = Student(
                student_id=str(row['Student ID']),
                course_enrolled=row['AC/IT'],
                attendance=row['Att'],
                last_name=row['Last Name'],
                first_name=row['First Name'],
                preferred_name=row['Preferred name']
            )
            session.add(student)

    session.commit()  # Commit the changes to the database
    return "Students added successfully"

def add_class_history(metadata, df):
    parsed_date = datetime.strptime(metadata['date'], '%m/%d/%Y').date()
    # Loop through the rows of your DataFrame and insert or update each row in the 'student_class' table
    for index, row in df.iterrows():
        # Check if a Student with the given student_id exists in the database
        existing_student = session.query(Student).filter_by(student_id=str(row['Student ID'])).first()
        
        if existing_student:
            # Now, create or update the StudentClass object
            existing_class_history = session.query(StudentClass).filter_by(student_id=str(row['Student ID']), class_name=metadata['class'], trainer_name=metadata['trainer']).first()
            if not existing_class_history:
                student_class = StudentClass(
                    student_id=str(row['Student ID']),
                    notes=row['Notes'],
                    date=parsed_date,
                    class_name=metadata['class'],
                    trainer_name=metadata['trainer'],
                    module_name=metadata['module'],
                    morning_break_status=row['9:00 AM - 11:00 AM'],
                    morning_break_hours=row['Morning Break\n(11:00 AM- 11:15 AM)'],
                    lunch_break_status=row['11:15 AM - 1:15 PM'],
                    lunch_break_hours=row['Lunch Break\n(1:15 PM -1:45 PM)'],
                    tea_break_status=row['1:45 PM - 3:45 PM'],
                    tea_break_hours=row['Tea Break\n(3:45 PM- 4:00 PM)'],
                    final_session_status=row['4:00 PM - 5:30 PM'],
                    final_session_hours=row['No of Hours at Final Session(5:30PM)'],
                    total_class_hours=row['Total Hours']
                )
                
                # Add the StudentClass to the session and commit the changes
                session.add(student_class)
                session.commit()

    return "Class History Added Successfully"


def calculate_break_hours(status):
    if status == "Present":
        return 2.0
    elif status == "Absent":
        return 0.0
    elif status == "Late_15_mins":
        return 1.75
    elif status == "Late_30_mins":
        return 1.5
    elif status == "Late_45_mins":
        return 1.25
    elif status == "Late_60_mins":
        return 1.0
    else:
        # Handle unknown status
        return 0.0  #